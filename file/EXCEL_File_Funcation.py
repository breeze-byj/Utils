# # 操作excel file的常用方法
import openpyxl


class ExcelFile:
    """
    封装一个处理excel文件的工具类
    """

    def __init__(self, path):
        """
        加载文件, 创建一个workbook对象
        :param path, excel文件路径
        """
        self.workbook = openpyxl.load_workbook(path)
        self.worksheet = None

    def Enca_get_sheet(self, var):
        """
        通过索引或者名字获取sheet
        :param var: 索引or名字
        :return: sheet
        """
        if isinstance(var, int):
            self.worksheet = self.workbook[self.workbook.sheetnames[var]]
            return self.worksheet

        if isinstance(var, str):
            self.worksheet = self.workbook[var]
            return self.worksheet

    def Enca_get_cell_value(self, row, col):
        """
        获取cell的值
        :param row: cell所在行
        :param col: cell所在列
        :return: cell的值
        """
        try:
            return self.worksheet.cell(row=row, column=col).value
        except BaseException as e:
            return None

    def Enca_get_max_row(self):
        """
        获取最大行数
        :return: 最大行数
        """
        return self.worksheet.max_row

    def Enca_get_min_row(self):
        """
        获取最小行数
        :return: 最小行数
        """
        return self.worksheet.min_row

    def Enca_get_max_col(self):
        """
        获取最大列数
        :return: 最大列数
        """
        return self.worksheet.max_column

    def Enca_get_min_col(self):
        """
        获取最小列数
        :return: 最小列数
        """
        return self.worksheet.min_column

    def Enca_get_all_value(self):
        """
        获取所有数据
        :return:
        """
        return tuple(self.worksheet.values)

    def Enca_get_row_data(self, row):
        """
        获取指定行的所有数据
        :param row: 指定行
        :return: 所有行数据组成的列表
        """
        return list(self.Enca_get_all_value()[row])

    def Enca_get_col_data(self, col):
        """
        获取指定列的所有数据
        :param col: 指定列, 如'A'
        :return: 所有列数据组成的列表
        """
        col_data_list = []
        for i in self.worksheet[col]:
            col_data_list.append(i.value)
        return col_data_list

    def Enca_write_data(self, row, col, value, path):
        """
        写入数据
        :param row: 指定cell所在行
        :param col: 指定cell所在列
        :param value: cell的值
        :param path: 保存文件路径
        """
        try:
            self.worksheet = self.workbook.active
            self.worksheet.cell(column=col, row=row, value=value)
            self.workbook.save(path)
        except BaseException as e:
            print(e)
            return None


if __name__ == '__main__':
    path = './test_case.xlsx'
    workbook = ExcelFile(path)

    workbook.Enca_get_sheet(1)
    print(workbook.Enca_get_sheet(1))
    # >>> <Worksheet "Sheet2">

    # 获取指定sheet
    workbook.Enca_get_sheet('Sheet1')
    print(workbook.Enca_get_sheet('Sheet1'))
    # >>> <Worksheet "Sheet1">

    # 获取指定cell的值
    data = workbook.Enca_get_cell_value(1, 3)
    print("data:", data)

    # 获取当前sheet某一行、列的数据
    data = workbook.Enca_get_row_data(2)
    print('第2行的数据为:', data)

    data = workbook.Enca_get_col_data('C')
    print('第c列的数据为:', data)

    # 写入数据
    workbook.Enca_write_data(11, 2, 'excel dispose', path)
